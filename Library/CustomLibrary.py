from docx import Document
import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import requests
import win32com.client as win32
import os
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import allure
import subprocess
from PyPDF2 import PdfReader
import pygetwindow as gw
import pyperclip
import pandas as pd


class CustomLibrary(object):

        def __init__(self):
                pass
        @property
        def _sel_lib(self):
            return BuiltIn().get_library_instance('SeleniumLibrary')

        @property
        def _driver(self):
            return self._sel_lib.driver

        def open_chrome_browser(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            try:
                options = webdriver.ChromeOptions()
                options.add_argument('--disable-gpu')
                options.add_argument("disable-extensions")
                options.add_argument('--ignore-ssl-errors=yes')
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--use-fake-ui-for-media-stream')
                options.add_experimental_option('prefs', {
                    'credentials_enable_service': False,
                    'profile': {
                        'password_manager_enabled': False
                    }
                })
                options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
                selenium.create_webdriver('Chrome',chrome_options=options)
                selenium.go_to(url)
                return True
            except:
                return False

        def create_dictionary_from_two_lists(self,key_list,value_list):
            # using dict() and zip() to convert lists to dictionary
            res = dict(zip(key_list, value_list))
            return  res

        def open_file(self, path):
            os.system(path)

        def print_screen(self):
            pyautogui.FAILSAFE = False
            pyautogui.keyDown("printscreen")
            pyautogui.keyUp("printscreen")
            time.sleep(2)

        def open_headless_chrome(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            options = webdriver.ChromeOptions()
            options.add_argument("--window-size=1440,900")
            options.add_argument('--disable-gpu')
            options.add_argument("disable-extensions")
            options.add_argument("--headless")
            options.add_experimental_option('prefs', {
                'credentials_enable_service': False,
                'profile': {
                    'password_manager_enabled': False
                }
            })
            options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
            selenium.create_webdriver('Chrome',chrome_options=options)
            selenium.go_to(url)
        
        def click_using_javascript(self,locator):
            element = self._sel_lib.get_webelement(locator)
            self._driver.execute_script("arguments[0].click();", element)
        
        def javascript_click(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script("arguments[0].click();", element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
            
        def get_text_by_using_javascript(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator) 
                return self._driver.execute_script("return arguments[0].textContent;", element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def wait_until_time(self,arg):
                time.sleep(int(arg))
            
        def wait_until_element_clickable(self,locator):
            try:
                """ An Expectation for checking that an element is either invisible or not present on the DOM."""
                if locator.startswith("//") or locator.startswith("(//"):
                    WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.XPATH, locator)))
                else:
                    WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.ID, locator)))
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        # def get_ms_excel_row_values_into_dictionary_based_on_key(self,filepath,keyName,sheetName=None):
        #     """Returns the dictionary of values given row in the MS Excel file """
            # workbook = xlrd.open_workbook(filepath)
            # snames = workbook.sheet_names()
            # dictVar = {}
            # if sheetName == None:
            #     sheetName = snames[0]      
            # if self.Verify_the_sheet_in_ms_excel_file(filepath,sheetName) == False:
            #     return dictVar
            # worksheet = workbook.sheet_by_name(sheetName)
            # noofrows = worksheet.nrows
            # dictVar = {}
            # headersList = worksheet.row_values(int(0))
            # for rowNo in range(1,int(noofrows)):
            #     rowValues = worksheet.row_values(int(rowNo))
            #     if str(rowValues[0])!= str(keyName):
            #         continue
            #     for rowIndex in range(0,len(rowValues)):
            #         cell_data = rowValues[rowIndex]
            #         if(str(cell_data) == "" or str(cell_data) == None):
            #             continue                    
            #         cell_data = self.get_unique_test_data(cell_data)
                
            #         dictVar[str(headersList[rowIndex])] = str(cell_data)
            # return dictVar 

        def get_ms_excel_row_values_into_dictionary_based_on_key(self, filepath, keyName, sheetName):
            """Returns the dictionary of values given row in the MS Excel file"""
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            dictVar = {}

            if sheetName is None:
                sheetName = snames[0]

            if sheetName not in snames or not self.Verify_the_sheet_in_ms_excel_file(filepath, sheetName):
                return dictVar

            worksheet = workbook[sheetName]
            headersList = [str(cell.value) for cell in worksheet[1]]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]) != str(keyName):
                    continue

                for rowIndex, cell_data in enumerate(row):
                    if cell_data is None or cell_data == "":
                        continue

                    # Ensure that get_unique_test_data is implemented correctly
                    cell_data = self.get_unique_test_data(cell_data)

                    dictVar[str(headersList[rowIndex])] = str(cell_data)

            return dictVar                 

        def get_unique_test_data(self,testdata):
            """Returns the unique if data contains unique word """
            ts = time.strftime("%H%M%S")
            unique_string = str(ts)
            testdata = testdata.replace("UNIQUE",unique_string)
            testdata = testdata.replace("Unique",unique_string)
            testdata = testdata.replace("unique",unique_string)
            return testdata

        def Verify_the_sheet_in_ms_excel_file(self,filepath,sheetName):
            """Returns the True if the specified work sheets exist in the specifed MS Excel file else False"""
            # workbook = xlrd.open_workbook(filepath)
            # snames = workbook.sheet_names()
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            sStatus = False        
            if sheetName == None:
                return True
            else:
                for sname in snames:
                    if sname.lower() == sheetName.lower():
                        wsname = sname
                        sStatus = True
                        break
                if sStatus == False:
                    print ("Error: The specified sheet: "+str(sheetName)+" doesn't exist in the specified file: " +str(filepath))
            return sStatus
        
        def clear_text_field(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script('arguments[0].value = "";', element)
            except Exception as e:
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def javascript_input_text(self,locator, text):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script('arguments[0].value = arguments[1];', element, text)
                self._driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", element)
                self._driver.execute_script('arguments[0].focus();', element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def download_PDF(self, project_folder, filename):
            pdf_url = self._driver.current_url
            response = requests.get(pdf_url)
            file_name = os.path.join(project_folder, filename)
            with open(file_name, 'wb') as f:
                f.write(response.content)
        
        def click_calendar_icon_in_vlms(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                # ActionChains(self._driver).move_by_offset(578, 465).click().perform()
                ActionChains(self._driver).move_to_element_with_offset(element, 67, 0).click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def click_element_with_offset(self, locator, x_divide, y_divide):
            try:
                element = self._sel_lib.get_webelement(locator)
                element_size = element.size
                print(element_size)
                print(float(x_divide))
                print(element_size['width'])
                offset_x = element_size['width'] // float(x_divide)
                offset_y = element_size['height'] // float(y_divide)
                print(offset_x, offset_y)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def right_click_element_with_offset(self, locator, x, y):
            try:
                element = self._sel_lib.get_webelement(locator)
                ActionChains(self._driver).move_to_element_with_offset(element, x, y).context_click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def switch_to_parent_frame(self):
            self._driver.switch_to.parent_frame()

        def screenshot_page(self,png_name):
            ul = BuiltIn().get_library_instance('SeleniumLibrary')
            path = ul.capture_page_screenshot()
            allure.attach.file(path, name=png_name, attachment_type=allure.attachment_type.JPG)
            return path

        def open_file_and_take_screenshot(self, path, file_name, kill='None', app='None'):
            subprocess.Popen([path], shell=True)
            time.sleep(7)
            screenshot = pyautogui.screenshot()
            cur_time = time.strftime("%H%M%S")
            file = file_name + cur_time+'.png'
            screenshot.save(file)
            if kill!='None': os.system("taskkill /f /im "+ app +".exe")
               
        def press_keyboard_key(self, key_name):
            try:
                pyautogui.press(key_name)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
 
        def press_multiple_keyboard_keys(self, key_name1, key_name2, key_name3):
            try:
                pyautogui.hotkey(key_name1, key_name2, key_name3)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def press_keyboard_keypairs(self, key_name1, key_name2):
            try:
                pyautogui.hotkey(key_name1, key_name2)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def input_text_with_offset(self, locator, x_divide, y_divide, text):
            try:
                element = self._sel_lib.get_webelement(locator)
                element_size = element.size
                print(element_size)
                print(float(x_divide))
                print(element_size['width'])
                offset_x = element_size['width']/float(x_divide)
                offset_y = element_size['height']/float(y_divide)
                print(offset_x, offset_y)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).click().send_keys(text).perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def double_click_element_with_offset(self, locator, x_divide, y_divide):
            try:
                element = self._sel_lib.get_webelement(locator)
                element_size = element.size
                print(element_size)
                print(float(x_divide))
                print(element_size['width'])
                offset_x = element_size['width']/float(x_divide)
                offset_y = element_size['height']/float(y_divide)
                print(offset_x, offset_y)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).double_click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def get_pdf_content(self, pdf_path):
            with open(pdf_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                pdf_content = ''
                for page_num in range(len(pdf_reader.pages)):
                    pdf_content += pdf_reader.pages[page_num].extract_text()
            return pdf_content
        
        def get_fullscreen_size(self):
            resolution = pyautogui.size()
            print(resolution)
        
        def get_word_window_size(self):
            word_window = gw.getWindowsWithTitle("Exe-8")
            if word_window:
                return word_window[0].size
            else:
                return None

        def get_chrome_url(self):
            chrome_window = gw.getActiveWindow()
            if chrome_window:
                # Focus on the Chrome window
                chrome_window.activate()
                # Send Ctrl + L (to focus on the address bar)
                pyautogui.hotkey('ctrl', 'l')
                time.sleep(1)  # Add a delay to ensure the address bar is focused
                # Send Ctrl + C (to copy the URL)
                pyautogui.hotkey('ctrl', 'c')
                time.sleep(1)  # Add a delay to ensure the copying is completed
                # Get the URL from the clipboard using pyperclip
                url = pyperclip.paste()
                pyautogui.hotkey('ctrl', 'w')
                return url.strip()  # Trim any leading/trailing spaces
            else:
                print("Chrome window not found.")
                return None

        def open_word_file(self, word_path, excel_file_path, sheet_name):
            try:
                # Call get_data_values to retrieve values from the Excel file
                data_values = self.get_data_values(excel_file_path, sheet_name)

                for key in data_values:
                    # Open the Word file using the file path in ColumnA
                    path = word_path+f'\\{key}'
                    subprocess.Popen(path, shell=True)
                    time.sleep(7)
                    # pyautogui.press('enter')
                    # time.sleep(3)
                    # pyautogui.keyDown('ctrl')
                    # pyautogui.keyDown('s')
                    # pyautogui.keyUp('s')
                    # pyautogui.keyUp('ctrl')
                    # time.sleep(2)
                    # pyautogui.press('e')
                    # time.sleep(2)
                    
                    for key1 in range(len(data_values[key]['ColumnB'])):
                        # To Find the value
                        time.sleep(1)
                        pyautogui.hotkey('ctrl', 'f')
                        time.sleep(2)
                        pyautogui.typewrite(data_values[key]['ColumnB'][key1])
                        pyautogui.press('enter')
                        time.sleep(2)
                        pyautogui.moveTo(1200, 420)
                        time.sleep(2)
                        pyautogui.keyDown('ctrl')
                        pyautogui.click()
                        pyautogui.keyUp('ctrl')
                        time.sleep(1)
                        pyautogui.press('enter')
                        time.sleep(2)
                        actual_url = self.get_chrome_url()
                        
                        if actual_url.lower() == data_values[key]['ColumnC'][key1].lower():
                            print(f"URL verification passed. Doc: {key} Actual URL: {actual_url}")
                        else:
                            print(f"URL verification failed. Doc: {key} Actual URL: {actual_url}, Expected URL: {data_values[key]['ColumnC'][key1]}")
                        
                        pyautogui.hotkey('alt', 'tab')
                        time.sleep(2)
                            
                os.system("taskkill /f /im WINWORD.EXE")
            except Exception as e:
                os.system("taskkill /f /im WINWORD.EXE")

        def get_data_values(self, file_path, sheet_name):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]

            data_values = {}

            for row in range(2, sheet.max_row + 1):
                # Get values from columns A, B, and C
                data = sheet.cell(row=row, column=1).value
                value_b = sheet.cell(row=row, column=2).value
                value_c = sheet.cell(row=row, column=3).value

                # If the data is not already in the dictionary, add it
                if data not in data_values:
                    data_values[data] = {'ColumnB': [], 'ColumnC': []}

                # Add values for the current data
                data_values[data]['ColumnB'].append(value_b)
                data_values[data]['ColumnC'].append(value_c)

            return data_values

        def get_column_values(self, file_path, sheet_name, column_name):
            try:
                # Read the Excel file
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Get the values of the specified column as a list
                column_values = df[column_name].tolist()

                return column_values

            except Exception as e:
                print(f"Error: {e}")
                return None
    
        def verify_response(self, project_folder='none', filename='none'):
                url = self._driver.current_url
                response = requests.get(url)
                # filename = time.strftime("%H%M%S")
                # file_name = os.path.join(project_folder, filename)
                # with open(file_name, 'wb') as f:
                #         f.write(response.content)
                if response.status_code == 200:
                    print(f"Success! The response code is {response.status_code}")
                else:
                    print(f"Error! The response code is {response.status_code}")
